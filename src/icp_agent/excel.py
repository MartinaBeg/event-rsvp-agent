"""Local Excel I/O for reading unprocessed rows and writing decisions in place."""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from .log import get_logger
from .models import Decision, RsvpRow

log = get_logger(__name__)

# Baseline schema column layout (1-indexed):
# A=Decision, B=Category, C=Priority, D=Role, E=Notes, F=#, G=Name, H=LinkedIn
COL_DECISION = 1
COL_CATEGORY = 2
COL_PRIORITY = 3
COL_ROLE = 4
COL_NOTES = 5
COL_ID = 6
COL_NAME = 7
COL_LINKEDIN = 8

# Input-layout fallback compatibility:
# A=Decision, B=Category, C=Priority, D=Notes, E=#, F=Name, G=LinkedIn
INPUT_COL_DECISION = 1
INPUT_COL_NAME = 6
INPUT_COL_LINKEDIN = 7

HEADER_ROWS = 1  # row 1 is the header; data starts at row 2
BASELINE_HEADERS = ["Decision", "Category", "Priority", "Role", "Notes", "#", "Name", "LinkedIn"]
DECISION_COLOR_FILLS: dict[str, PatternFill] = {
    "Approve": PatternFill(fill_type="solid", fgColor="FFC6EFCE"),
    "Waitlist": PatternFill(fill_type="solid", fgColor="FFFFEB9C"),
    "Reject": PatternFill(fill_type="solid", fgColor="FFFFC7CE"),
}
NOTES_WIDTH_EXPANSION_FACTOR = 1.5
COLUMN_MIN_WIDTHS: dict[int, float] = {
    COL_DECISION: 10,
    COL_CATEGORY: 14,
    COL_PRIORITY: 10,
    COL_ROLE: 22,
    COL_NOTES: 28,
    COL_ID: 6,
    COL_NAME: 16,
    COL_LINKEDIN: 26,
}
COLUMN_MAX_WIDTHS: dict[int, float] = {
    COL_DECISION: 14,
    COL_CATEGORY: 24,
    COL_PRIORITY: 12,
    COL_ROLE: 38,
    COL_NOTES: 80,
    COL_ID: 8,
    COL_NAME: 28,
    COL_LINKEDIN: 40,
}


class WorkbookClient(Protocol):
    def read_unprocessed(self) -> list[RsvpRow]: ...
    def write_decisions(self, decisions: list[Decision]) -> None: ...


class ExcelClient:
    """openpyxl-backed reader/writer for a local .xlsx file."""

    def __init__(self, file_path: str | Path, tab_name: str = "RSVP List") -> None:
        self._path = Path(file_path).expanduser().resolve()
        self._tab_name = tab_name
        if not self._path.is_file():
            raise FileNotFoundError(f"Excel file not found: {self._path}")

    def read_unprocessed(self) -> list[RsvpRow]:
        """Return rows where LinkedIn has a URL and Decision is empty."""
        wb = load_workbook(self._path, data_only=True)
        if self._tab_name not in wb.sheetnames:
            raise ValueError(
                f"Tab {self._tab_name!r} not found in workbook. "
                f"Available: {wb.sheetnames}"
            )
        ws = wb[self._tab_name]
        cols = _resolve_read_columns(ws)

        rows: list[RsvpRow] = []
        total_data_rows = 0
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True), start=HEADER_ROWS + 1
        ):
            row = list(row) + [None] * max(cols["linkedin"] - len(row), 0)
            decision = _stringy(row[cols["decision"] - 1])
            url = _stringy(row[cols["linkedin"] - 1])
            name = _stringy(row[cols["name"] - 1])
            if not url:
                continue
            total_data_rows += 1
            if decision:
                continue
            rows.append(
                RsvpRow(
                    row_number=row_idx,
                    name=name or "(unknown)",
                    linkedin_url=url,
                )
            )
        wb.close()
        log.info(
            "excel.read",
            path=str(self._path),
            tab=self._tab_name,
            unprocessed=len(rows),
            total_with_url=total_data_rows,
        )
        return rows

    def write_decisions(self, decisions: list[Decision]) -> None:
        """Write decisions with baseline schema columns and save in place."""
        if not decisions:
            log.info("excel.write_skipped", reason="no_decisions")
            return
        wb = load_workbook(self._path)
        if self._tab_name not in wb.sheetnames:
            raise ValueError(
                f"Tab {self._tab_name!r} not found in workbook. "
                f"Available: {wb.sheetnames}"
            )
        ws = wb[self._tab_name]
        _ensure_baseline_schema(ws)
        for d in decisions:
            decision_cell = ws.cell(row=d.row_number, column=COL_DECISION, value=d.decision)
            fill = DECISION_COLOR_FILLS.get(d.decision)
            if fill is not None:
                decision_cell.fill = fill
            ws.cell(row=d.row_number, column=COL_CATEGORY, value=d.category)
            ws.cell(row=d.row_number, column=COL_PRIORITY, value=d.priority)
            role_value = _stringy(d.role) or "N/A"
            notes_value = _stringy(d.notes) or f"{role_value} — profile reviewed"
            ws.cell(row=d.row_number, column=COL_ROLE, value=role_value)
            ws.cell(row=d.row_number, column=COL_NOTES, value=notes_value)
        _apply_linkedin_hyperlinks(ws)
        _adjust_column_widths(ws)
        wb.save(self._path)
        wb.close()
        log.info("excel.write", path=str(self._path), rows=len(decisions))


def _stringy(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _header_map(ws: object) -> dict[str, int]:
    out: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = _stringy(ws.cell(row=1, column=col).value).lower()
        if key and key not in out:
            out[key] = col
    return out


def _resolve_read_columns(ws: object) -> dict[str, int]:
    headers = _header_map(ws)
    return {
        "decision": headers.get("decision", INPUT_COL_DECISION),
        "name": headers.get("name", INPUT_COL_NAME),
        "linkedin": headers.get("linkedin", INPUT_COL_LINKEDIN),
    }


def _ensure_baseline_schema(ws: object) -> None:
    headers = _header_map(ws)
    # If the sheet has LinkedIn in col G and no Role column, shift to baseline.
    if "role" not in headers and headers.get("linkedin", 0) == INPUT_COL_LINKEDIN:
        ws.insert_cols(COL_ROLE)
    for i, header in enumerate(BASELINE_HEADERS, start=1):
        ws.cell(row=1, column=i, value=header)
    _ensure_role_column_style(ws)


def _ensure_role_column_style(ws: object) -> None:
    """Keep Role column styling consistent with the Notes template column."""
    max_row = max(ws.max_row, HEADER_ROWS)
    for row in range(1, max_row + 1):
        source = ws.cell(row=row, column=COL_NOTES)
        target = ws.cell(row=row, column=COL_ROLE)
        if source.has_style:
            target._style = copy(source._style)

    notes_letter = get_column_letter(COL_NOTES)
    role_letter = get_column_letter(COL_ROLE)
    notes_width = ws.column_dimensions[notes_letter].width
    if notes_width is not None:
        ws.column_dimensions[role_letter].width = notes_width


def _apply_linkedin_hyperlinks(ws: object) -> None:
    max_row = max(ws.max_row, HEADER_ROWS)
    for row in range(HEADER_ROWS + 1, max_row + 1):
        cell = ws.cell(row=row, column=COL_LINKEDIN)
        url = _stringy(cell.value)
        if not url:
            continue
        normalized = _normalize_linkedin_url(url)
        if not normalized:
            continue
        cell.hyperlink = normalized


def _normalize_linkedin_url(url: str) -> str:
    value = url.strip()
    lower = value.lower()
    if "linkedin.com/" not in lower:
        return ""
    if lower.startswith("http://") or lower.startswith("https://"):
        return value
    return f"https://{value.lstrip('/')}"


def _adjust_column_widths(ws: object) -> None:
    max_row = max(ws.max_row, HEADER_ROWS)
    width_by_col: dict[int, float] = {}
    for col in range(COL_DECISION, COL_LINKEDIN + 1):
        max_len = 0
        for row in range(1, max_row + 1):
            value = _stringy(ws.cell(row=row, column=col).value)
            if len(value) > max_len:
                max_len = len(value)
        min_w = COLUMN_MIN_WIDTHS.get(col, 8.0)
        max_w = COLUMN_MAX_WIDTHS.get(col, 50.0)
        width = max(min_w, min(max_w, max_len + 2))
        width_by_col[col] = width

    notes_width = width_by_col.get(COL_NOTES, COLUMN_MIN_WIDTHS[COL_NOTES])
    notes_max = COLUMN_MAX_WIDTHS[COL_NOTES]
    width_by_col[COL_NOTES] = min(notes_max, max(notes_width, notes_width * NOTES_WIDTH_EXPANSION_FACTOR))

    for col, width in width_by_col.items():
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = float(width)
