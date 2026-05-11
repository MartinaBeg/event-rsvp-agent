"""Round-trip test for the Excel reader/writer using a synthetic workbook.

Avoids any reliance on the take-home example file so this runs cleanly in CI.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from icp_agent.excel import ExcelClient
from icp_agent.models import Decision


def _make_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RSVP List"
    ws.append(["Decision", "Category", "Priority", "Notes (optional)", "#", "Name", "LinkedIn"])
    ws.append([None, None, None, None, 1, "Alice", "https://linkedin.com/in/alice"])
    ws.append([None, None, None, None, 2, "Bob", "https://linkedin.com/in/bob"])
    # Pre-decided row should be skipped on read.
    ws.append(["Approve", "Engineer", "P1", "already done", 3, "Carol", "https://linkedin.com/in/carol"])
    # Row with no URL is skipped entirely.
    ws.append([None, None, None, None, 4, "Dan", None])
    wb.save(path)


def test_excel_read_skips_processed_and_empty(tmp_path: Path) -> None:
    p = tmp_path / "rsvp.xlsx"
    _make_workbook(p)
    client = ExcelClient(p, tab_name="RSVP List")
    rows = client.read_unprocessed()
    assert [r.name for r in rows] == ["Alice", "Bob"]
    assert rows[0].row_number == 2
    assert rows[1].row_number == 3


def test_excel_write_round_trip(tmp_path: Path) -> None:
    p = tmp_path / "rsvp.xlsx"
    _make_workbook(p)
    client = ExcelClient(p, tab_name="RSVP List")

    decisions = [
        Decision(
            row_number=2,
            name="Alice",
            linkedin_url="https://linkedin.com/in/alice",
            category="Industry Researcher",
            priority="P1",
            decision="Approve",
            role="Research Scientist @ OpenAI",
            notes="Research Scientist @ OpenAI",
            score=8,
        ),
        Decision(
            row_number=3,
            name="Bob",
            linkedin_url="https://linkedin.com/in/bob",
            category="Other",
            priority="P5",
            decision="Reject",
            role="N/A",
            notes="Scrape failed — no rescuable signal",
            score=0,
        ),
    ]
    client.write_decisions(decisions)

    wb = load_workbook(p)
    ws = wb["RSVP List"]
    assert [ws.cell(row=1, column=c).value for c in range(1, 9)] == [
        "Decision",
        "Category",
        "Priority",
        "Role",
        "Notes",
        "#",
        "Name",
        "LinkedIn",
    ]
    assert ws.cell(row=2, column=1).value == "Approve"
    assert ws.cell(row=2, column=2).value == "Industry Researcher"
    assert ws.cell(row=2, column=3).value == "P1"
    assert ws.cell(row=2, column=4).value == "Research Scientist @ OpenAI"
    assert ws.cell(row=2, column=5).value == "Research Scientist @ OpenAI"
    assert ws.cell(row=3, column=1).value == "Reject"
    # Pre-decided row 4 should be untouched.
    assert ws.cell(row=4, column=1).value == "Approve"
    assert ws.cell(row=4, column=5).value == "already done"


def test_excel_idempotent(tmp_path: Path) -> None:
    """Re-reading after writing should yield no unprocessed rows."""
    p = tmp_path / "rsvp.xlsx"
    _make_workbook(p)
    client = ExcelClient(p, tab_name="RSVP List")
    decisions = [
        Decision(
            row_number=2,
            name="Alice",
            linkedin_url="https://linkedin.com/in/alice",
            category="Engineer",
            priority="P3",
            decision="Approve",
            notes="ok",
            score=4,
        ),
        Decision(
            row_number=3,
            name="Bob",
            linkedin_url="https://linkedin.com/in/bob",
            category="Engineer",
            priority="P5",
            decision="Reject",
            notes="ok",
            score=0,
        ),
    ]
    client.write_decisions(decisions)
    assert client.read_unprocessed() == []
